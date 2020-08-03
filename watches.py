from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
from dataclasses import dataclass
import requests
import time
import os.path
import shutil
from openpyxl import Workbook,load_workbook

@dataclass
class Brand:
    label: str
    link: str
    watchcount: int

@dataclass
class Watch:
    brand: str
    name: str
    price: str
    product_emi: str
    url: str
    id: str
    image_src: str
    image_filename: str


class WatchScraper:
    def __init__(self, ):
        self.base_url = "https://www.ethoswatches.com/brands.html"
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.implicitly_wait(50)  # Set impicit waits for 50 seconds
        self.brand_list = []
        self.watch_data_list = []
        self.target_path = "./images/"
        self.filename = "watch.xlsx"
        self.MAX_WatchesPerBrand = 2


    def get_brands(self):
        soup = BeautifulSoup(self.driver.page_source, 'html.parser')
        brands = soup.find(class_='items am_shopby_filter_items_attr_manufacturer')
        brands = brands.find_all(class_='item')

        for brand in brands:
            link = brand.find('a').get('href')
            label = brand.get_text()
            label_count_list = label[3:].split('\n')
            label = label_count_list[0]
            watchcount = int(label_count_list[1])
            self.brand_list.append(Brand(label, link, watchcount))


    def get_productlist(self,brandlink):
        self.driver.get(brandlink)
        print(f"Brand url - {brandlink}, Get products...")
        soup = BeautifulSoup(self.driver.page_source, 'html.parser')
        watchlist1 = soup.find(class_='columns')
        watchlist_soup = watchlist1.div.find_all('li')
        product_list = []
        for watch in watchlist_soup:
            if watch.find('h2') is not None:
                link = watch.find('a').get('href')
                product_list.append(link)
        return product_list

    def download_image(self, src, filename):
        #Download Images to a local directory
        try:
            response = requests.get(src, stream=True)
            file_path = os.path.join(self.target_path, filename)
            if not os.path.isdir(self.target_path):
                os.mkdir(self.target_path)
            file = open(file_path, 'wb')
            response.raw.decode_content = True
            shutil.copyfileobj(response.raw, file)
            del response
            print("Product Image downloaded...")
            return True
        except Exception as e:
            print(e)
            return False

    def createExcelfile(self):
        heading = ["Brand", "Name", "id", "Price", "EMI", "ImageFileName"]
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(heading)
        workbook.save(filename=self.filename)
        workbook.close()

    def saveto_excel(self):
        print("Save to Excel")
        workbook = load_workbook(filename=self.filename)
        sheet = workbook.active
        for watch in self.watch_data_list:
            rowdata = [watch.brand, watch.name, watch.id, watch.price, watch.product_emi, watch.image_filename]
            sheet.append(rowdata)
            workbook.save(filename=self.filename)
        workbook.close()

    def get_watchdetails(self,product_list,brandlabel):
        watchcount_init = 0
        self.watch_data_list = []
        for product_link in product_list:
            if watchcount_init >= self.MAX_WatchesPerBrand: break
            self.driver.get(product_link)
            soup = BeautifulSoup(self.driver.page_source, 'html.parser')
            product_info = soup.find(class_='product_info_main')
            brand = brandlabel
            name = product_info.h1.span.text.strip()
            id = product_info.h3.text.strip()
            price = product_info.find(class_='price').text
            product_emi = product_info.find(class_='product_emi')
            if product_emi is not None:
                product_emi = product_emi.text.strip()
            image_info = soup.find(class_='openPhotoSwipe').findChildren("img")
            image_info_tag = image_info[0]["alt"]
            image_info_src = image_info[0]["src"]
            realname = ''.join(e for e in image_info_tag if e.isalnum())
            image_filename = realname + str(id) + ".jpg"
            image_filename = image_filename.replace("/", "-")

            if self.download_image(image_info_src, image_filename):
                self.watch_data_list.append(Watch(brand, name, price, product_emi, product_link, id, image_info_src, image_filename))
                watchcount_init += 1


    def main_logic(self):
        self.createExcelfile()
        self.driver.get(self.base_url)
        self.driver.maximize_window()
        self.get_brands()
        for brand in self.brand_list:
            time.sleep(1)
            product_list = self.get_productlist(brand.link)
            self.get_watchdetails(product_list, brand.label)
            self.saveto_excel()

        time.sleep(3)
        self.driver.quit()

if __name__ == '__main__':
    watch_details = WatchScraper()
    watch_details.main_logic()